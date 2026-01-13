import openpyxl
import os
import time
import pandas as pd



def main():
    print("这里是检测单词excel的文件，请输入你想操作的语言")
    print("直接回车.日语")
    print("1.法语")
    print("2.英语")
    root_path = r"D:\OneDrive\OneSyncFiles\data"
    lang = _input_params(["","1","2"])
    if lang == "" :
        path = os.path.join(root_path,"japanese.xlsx")
    elif lang == "1":
        path = os.path.join(root_path, "france.xlsx")
    elif lang == "2":
        path = os.path.join(root_path, "english.xlsx")

    print("当前的excel文件")
    print(path)


    print("请输入你想做的事儿：")
    print("直接回车.查看这个excel里每个sheet的行数 以及 查看Q列所有重复的词汇")
    print("1.查看考试的单词数")
    select = _input_digital_or_none()
    if select == "":
        count_rows_in_excel(path)
        print()
        print("++++++++++++++++++++++++++++++++++++")
        print()
        column = "Q"
        ZhengGeExcelChongFuCiHui(path,column)
    elif select == "1":
        count_value_ratio_in_all_sheets(path,"T","N")
    else:
        print("输入错误")

    input()


def count_value_ratio_in_all_sheets(path, column_name, keyword):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(path)

    test_count = 0
    total_count = 0
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_total_count = sheet.max_row - 1
        N_count = _count_n_in_column_q(sheet,column_name,keyword)
        keyword_count = sheet_total_count - N_count
        print(f"Sheet '{sheet_name}': {keyword_count}/{sheet_total_count}")
        test_count += keyword_count
        total_count += sheet_total_count

    print("===================")
    print(f"总共：'{test_count}/{total_count}'")



# 这个方法的作用是，在某个sheet中，T列的，N这个值的统计
def _count_n_in_column_q(sheet,column_name,key_word):
    for col_index, cell in enumerate(sheet[1], start=1):  # 假设第一行为表头
        if cell.value == column_name:
            column_index = col_index
            break

    if column_index is None:
        raise ValueError(f"列名 '{column_name}' 不存在！")

        # 遍历列，统计关键字出现的次数
    count = 0
    for row in sheet.iter_rows(min_col=column_index, max_col=column_index, min_row=2, values_only=True):  # 从第二行开始
        if row[0] == key_word:
            count += 1

    return count

    return n_count


def ZhengGeExcelChongFuCiHui(path, column):
    # 读取所有 sheet
    sheets = pd.read_excel(path, sheet_name=None)

    # 构建初始字典，key 为 sheet 名，value 为指定列的所有数据
    sheet_data_dict = {}
    sheet_id_dict = {}

    for sheet_name, data in sheets.items():
        if column in data.columns and "id" in data.columns:
            sheet_data_dict[sheet_name] = data[column].dropna().tolist()
            sheet_id_dict[sheet_name] = data["id"].dropna().tolist()

    # 比较数据，构建结果字典
    comparison_dict = {}
    seen_pairs = set()  # 用于去重

    # 遍历字典的所有 key 和对应的值
    for sheet1, data1 in sheet_data_dict.items():
        # 比较不同 sheet 间的数据
        for sheet2, data2 in sheet_data_dict.items():
            if sheet1 != sheet2:  # 避免比较同一个 sheet
                for index1, value1 in enumerate(data1):
                    for index2, value2 in enumerate(data2):
                        if value1 == value2:  # 如果值相等
                            id1 = sheet_id_dict[sheet1][index1]
                            id2 = sheet_id_dict[sheet2][index2]
                            pair = tuple(sorted([(sheet1, id1, value1), (sheet2, id2, value2)]))
                            if pair not in seen_pairs:  # 检查是否已记录
                                key = f"{sheet1}_{id1}_{value1}"
                                value = f"{sheet2}_{id2}_{value2}"
                                comparison_dict[key] = value
                                seen_pairs.add(pair)

        # 比较本 sheet 内的数据
        for index1, value1 in enumerate(data1):
            for index2, value2 in enumerate(data1):
                if index1 != index2 and value1 == value2:  # 排除自身比较，寻找重复值
                    id1 = sheet_id_dict[sheet1][index1]
                    id2 = sheet_id_dict[sheet1][index2]
                    pair = tuple(sorted([(sheet1, id1, value1), (sheet1, id2, value2)]))
                    if pair not in seen_pairs:  # 检查是否已记录
                        key = f"{sheet1}_{id1}_{value1}"
                        value = f"{sheet1}_{id2}_{value2}"
                        comparison_dict[key] = value
                        seen_pairs.add(pair)

    # 打印结果字典
    print("Comparison Dictionary:")
    if len(comparison_dict) != 0:
        for k, v in comparison_dict.items():
            print(f"{k}: {v}")
            print("")
    else:
        print("没有！")

    return comparison_dict

def count_rows_in_excel(file_path):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 初始化总行数
    total_rows = 0
    sheet_row_counts = {}

    # 遍历每个工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        row_count = sheet.max_row - 1  # 获取当前工作表的最大行数(标题有一行)
        sheet_row_counts[sheet_name] = row_count
        total_rows += row_count  # 累加行数

    # 输出每个工作表的行数和总行数
    print("每个工作表的行数:")
    for sheet, rows in sheet_row_counts.items():
        print(f"{sheet}: {rows} 行")

    print(f"整个 Excel 文件的总行数: {total_rows} 行")
    return sheet_row_counts, total_rows



# 这个方法只能输入一个正确的file path
def _input_file():
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在 并且是一个file
        if os.path.exists(path) and os.path.isfile(path):
            # 存在并且是一个file的情况，返回此路径，循环结束
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的file路径!")
        time.sleep(1)
        print("请重新输入")

# 这个方法只能输入数字或者空
def _input_digital_or_none():
    while True:
        select = input()
        if select.isdigit() or select == "":
            return select
        print("请输入数字～～")
        time.sleep(1)
        print("请重新输入")

def _input_params(check_collection=[]):
    while True:
        select = input()
        if select in check_collection:
            return select
        print("输入的参数不匹配！")
        time.sleep(1)
        print("请重新输入")

if __name__ == '__main__':
    main()
    # path = r"D:\OneDrive\OneSyncFiles\data\japanese.xlsx"
    # column = "Q"
    # ZhengGeExcelChongFuCiHui(path,column)
