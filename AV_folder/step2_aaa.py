'''
    这个类是在step1执行成功后，才执行的
'''
from genericpath import isdir
from multiprocessing.util import is_exiting
import os,sys,re,subprocess
from pathlib import Path
import json,shutil

import openpyxl
from sympy import im, true
from torch import le

from openpyxl import Workbook


zone_folder_frame= 'zone_{no_str}'


def main():
    folder_path = r'H:\japan'
    folder_zon_name = 'zone_01'
    all_video_count,file_dict,folder_dict = _get_all_file_path(folder_path)

    folder_path = r'H:\japan_img'
    aaa(folder_path)



    # seperate_muilti_list_and_make_zone(file_dict,folder_path)

    # final_check_if_right(folder_path)

    # check_folder_all_in_zone1(folder_dict,folder_zon_name)

    # make_excel(folder_path)
    

    # img_folder = r'H:\japan_img\zone_01\ssss'
    # folder_img_move_folder(img_folder,folder_dict)

    # img_folder = r'H:\japan_img'
    # for folder_1 in os.listdir(img_folder):
    #     img_folder_1 = os.path.join(img_folder,folder_1)
    #     file_img_move_folder(img_folder_1,file_dict)

    # view_all_folder_video_unit(folder_dict)

def aaa(folder_path):
    for folder in os.listdir(folder_path):
        cur_folder_path = os.path.join(folder_path,folder)
        if os.path.isdir(cur_folder_path) and folder.startswith('zone_'):
            for file in os.listdir(cur_folder_path):
                cur_file_path = os.path.join(cur_folder_path,file)
                print()
                print(cur_file_path)
                print(folder_path)
                shutil.move(cur_file_path,folder_path)

            


def seperate_muilti_list_and_make_zone(file_dict,folder_path):
    values = list(file_dict.values())
    # print(values)

    size = 18
    seperated_list = []
    for index in range(0,len(values),size):
        seperated_list.append(values[index:size+index])
    
    zone_no = 2
    for cur_list in seperated_list:
        no_str = _get_str_of_number(zone_no)

        cur_zone_folder_path = os.path.join(folder_path,zone_folder_frame.format(no_str=no_str))
        
        for cur_file in cur_list:
            # print(cur_file + '\t' + cur_zone_folder_path)
            print(os.path.basename(cur_file) + '\t' + os.path.basename(cur_zone_folder_path))
            os.makedirs(cur_zone_folder_path,exist_ok=True)
            shutil.move(cur_file,cur_zone_folder_path)

        zone_no += 1

    
def _get_str_of_number(no):
    if no < 10:
        return '0' + str(no)
    else:
        return str(no)


def check_folder_all_in_zone1(folder_dict,folder_zone_name):
    '''
        是否所有的文件夹分段视频都在zone_1里？
    '''
    for key in folder_dict.keys():
        cur_folder_path = folder_dict[key]
        if not folder_zone_name in cur_folder_path:
            print(f'居然有文件夹不在{folder_zone_name}里！')
            print(cur_folder_path)
            sys.exit(1)
    
    print(f'验证完成，所有的文件夹分段视频都在{folder_zone_name}里')



def final_check_if_right(folder_path):
    '''
        最终的excel和视频文件是否对应的验证
    '''
    excel_path = os.path.join(os.path.dirname(folder_path),'J_info.xlsx')
    book = openpyxl.load_workbook(excel_path,read_only=True)
    sheet = book['info']

    result = []

    # 如果有标题，从第二行开始：range(2, ws.max_row + 1)
    for row in range(2, sheet.max_row + 1):  
        b_val = sheet[f"B{row}"].value
        e_val = sheet[f"E{row}"].value

        row_dict = {
            "B": b_val,
            "E": e_val
        }
        result.append(row_dict)
    
    print('excel 里的数据获取完毕！现在开始验证')

    for dict_data in result:
        fild_path = os.path.join(folder_path,dict_data['E'],dict_data['B'])
        if not os.path.exists(fild_path):
            print(f'验证失败，这个文件{fild_path}居然不存在')
            sys.exit(1)
    
    print(f'验证成功！！{folder_path}里没有问题！')
    


def view_all_folder_video_unit(folder_dict):
    for key in folder_dict.keys():
        print(folder_dict[key])
    
    print(len(folder_dict))


def file_img_move_folder(img_folder,file_dict):
    '''
        img_folder
            所有的图片都放在这个folder里    
        
        file_dict
            参照着这个字典，来将图片分别放入各自的folder里

    '''


    # 获取上层的路径，因为这里默认img_folder 和其他zone folder 在同级
    output_root = os.path.dirname(img_folder)

    for img in os.listdir(img_folder):
        if not img.endswith('.png'):
            print('这路径底下，不应该有不是png的文件，你检查下')
            sys.exit(1)
        
        img_info_list = img.split('.')  
        video_name = img_info_list[0] + '.' + img_info_list[1]
        if video_name in file_dict:
            original_path = file_dict[video_name]
            folder_namne = os.path.basename(os.path.dirname(original_path))
            folder_path = os.path.join(output_root,folder_namne)
            cur_img_path = os.path.join(img_folder,img)
            print()
            print(cur_img_path)
            print(folder_path)
            os.makedirs(folder_path,exist_ok=True)

            # 新路径（保持原文件名）
            new_path = os.path.join(folder_path, os.path.basename(cur_img_path))
            # 移动文件
            shutil.move(cur_img_path, new_path)

        else:
            sys.exit(1)
        

def folder_img_move_folder(img_folder,folder_dict):
    '''
        img_folder
            这里底下只有img的文件    
        
        folder_dict
            就是_get_all_file_path的返回值的folder_dict
    '''

    for folder_path in list(folder_dict.values()):

        new_img_folder = folder_path.replace('japan','japan_img')
        os.makedirs(new_img_folder,exist_ok=True)

        for file in os.listdir(folder_path):
            
            file_check = os.path.basename(file).split('.')[0]

            for img in os.listdir(img_folder):
                img_basename = img.split('.')[0]
                if file_check == img_basename:
                    old_img_path = os.path.join(img_folder,img)
                    print(old_img_path)
                    print(new_img_folder)
                    shutil.move(old_img_path,new_img_folder)


def _get_all_file_path(folder_path):
    '''
        获取所有文件的路径，默认为从下面的路径中获取
            root dir
                zone dir
                    file/folder(一个视频文件分段的)
        
        获取后的格式，例：
            file_list: {jul-034.mp4,c://XXXX//XXX//jul-034.mp4}
            folder_list: {jul-035,c://XXXX//XXX//jul-034}
        
    '''

    # 初始化
    # 所有视频文件个数(包括文件夹分段的视频)
    all_video_count = 0
    # 所有单个视频文件
    file_dict = {}
    # 所有分段的文件夹视频
    folder_dict = {}

    # 获取所有文件的路径
    for zone_folder_name in os.listdir(folder_path):
        zone_folder_path = os.path.join(folder_path,zone_folder_name)

        # root文件夹下，只能有各个分区的文件夹，不能有文件，root文件夹下，只要有文件存在，就shut down
        # 各个分区的文件夹下，才是视频
        if not os.path.isdir(zone_folder_path):
            print(zone_folder_path)
            print('root路径下，居然有不是folder 的文件！程序结果，你自己检查下吧')
            sys.exit(1)
        
        # 遍历分区文件夹下的所有文件 这里应该每个都是视频文件(或者分段视频文件夹)
        for video_file_path in os.listdir(zone_folder_path):

            all_video_count += 1
            cur_file = os.path.join(zone_folder_path,video_file_path)

            if os.path.isfile(cur_file):
                file_dict[video_file_path] = cur_file
            else:
                folder_dict[video_file_path] = cur_file
    
    return all_video_count,file_dict,folder_dict


def make_excel(folder_path):
    
    output_excel = r'H:\J_info.xlsx'

    all_data = {}
    # check 文件名开头是否是字母
    for folder in os.listdir(folder_path):
        if not re.fullmatch(r"[a-z]+",folder[0]):
            print(folder)
        else:
            cur_folder = os.path.join(folder_path,folder)
            for video_file in os.listdir(cur_folder):
                all_data[video_file] = folder[1:]


    data_list = []
    # 打印 
    for key in all_data.keys():
        is_folder = False
        cur_data = []
        cur_data.append(key)
        cur_data.append(all_data[key])
        if key.count('.') == 0:
            cur_data.append('folder')
            is_folder = True
        else:
            cur_data.append('file')
        
        # if is_folder:
        #     data_list.append(cur_data)

        data_list.append(cur_data)
    
    
    # check 出力内容
    print(len(data_list))
    for data in data_list:
        print(data)
    
    # _list_to_excel(data_list,output_excel)
    

def _list_to_excel(data_list, output_path):
    wb = Workbook()
    ws = wb['info']

    for row in data_list:
        ws.append(row)

    wb.save(output_path)
    print("Excel 已生成:", output_path)



if __name__ == '__main__':
    main()