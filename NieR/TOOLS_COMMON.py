import datetime, platform, socket, threading, subprocess, os, sys, time
import shutil
from decimal import Decimal
from tkinter import NO
from turtle import st
import cv2, numpy, librosa
from docx2pdf import convert
import PyPDF2
from openpyxl.styles import Font
import TOOLS_COMMON as common

RETRY_TIMES = 5

def none_or_what(default_value):
    input_str = input()
    if input_str == "":
        print("default_value++++")
        input_str = default_value
    return input_str


def get_desktop_path():
    system = platform.system()
    if system == "Windows":
        return os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    elif system == "Darwin":  # macOS
        return os.path.join(os.path.expanduser('~'), 'Desktop')
    else:  # Linux
        return os.path.join(os.path.expanduser('~'), 'Desktop')

# 一定要在刚好save之前使用
def set_excel_font_style(sheet,style = "微软雅黑", size = 16):
    # 定义全局字体
    default_font = Font(name=style, size=size)
    # 设置全局字体样式
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = default_font


# 删除一个a文件夹中除b文件名以外的所有文件
def delete_all_file_except(folder_path,except_file_name):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        # 检查当前文件是否在保留列表中
        if os.path.basename(filename) != except_file_name:
            print()
            # 如果不在保留列表中，则删除
            try:
                # 检查是否为文件，是则删除
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                # 检查是否为文件夹，是则递归删除该文件夹及其内容
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f'删除 {file_path} 时发生错误。错误信息: {e}')

# 入力：1.docx文件的folder   2.合成后pdf的文件名
def merge_pdfs_design(folder_path,merged_file_name):
    merged_file = os.path.join(folder_path,merged_file_name)

    files = []

    for file in os.listdir(folder_path):
        if file.endswith(".pdf"):
            cur_file = os.path.join(folder_path, file)
            files.append(cur_file)

    # print(files)
    print("====================")
    # print(merged_file)
    print("开始合成！")
    merge_pdfs(files, merged_file)
    print("合成完毕！")


# 入力：1.全是pdf全路径文件的list      2.合成后的pdf 路径
def merge_pdfs(pdf_files, output_pdf):
    pdf_writer = PyPDF2.PdfWriter()

    for path in pdf_files:
        pdf_reader = PyPDF2.PdfReader(path)
        for page in range(len(pdf_reader.pages)):
            # 将每页添加到writer对象
            pdf_writer.add_page(pdf_reader.pages[page])

    # 写出合并的PDF
    with open(output_pdf, 'wb') as out:
        pdf_writer.write(out)

# 将folder里所有的doc文件导出成pdf文件
def word_to_pdf(select_path):
    for file in os.listdir(select_path):
        if file.endswith(".docx"):
            cur_file_full_path = os.path.join(select_path,file)
            new_ext_file = switch_ext(cur_file_full_path,".pdf")
            print(cur_file_full_path)
            print(new_ext_file)
            convert(cur_file_full_path,new_ext_file)
            print("======================")
            print()

# 获取当前系统的查找指令
def get_find_cmd():
    os = get_os()
    if os == "w":
        return "findstr"
    elif os == "l":
        return "grep"

# 获取一个视频的fps，resolution和全部的frames数量
def get_fps_resolution_allFrames(video_path):
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)

    width = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
    height = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
    resolution = [width, height]

    frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)

    return fps, resolution, frames


# 获取peak_amplitudes
def get_peak_amplitudes(video_path):
    # 将给入的fps给四舍五入一下，获得新的fps
    fps = round_to_the_neares(get_fps_resolution_allFrames(video_path)[0])
    print("fps:", fps)

    # 使用librosa处理video file，通过获取的y和sr的对象以及某个算法，算出peak_amplitudes后返回
    y, sr = librosa.load(video_path)
    frame_size = sr // int(fps)
    num_frames = len(y) // frame_size
    peak_amplitudes = [numpy.max(numpy.abs(y[i * frame_size:(i + 1) * frame_size])) for i in range(int(num_frames))]

    return peak_amplitudes


# 四舍五入
def round_to_the_neares(num):
    return Decimal(num).quantize(Decimal("1."), rounding="ROUND_HALF_UP")


# 给我一个file路径，返回这个file名，如：入力：/home/aaa.txt，出力：aaa
def get_filename_only(file_path):
    return os.path.splitext(os.path.basename(file_path))[0]


# 入力：file full path，出力：只是换了新后缀；不给new_ext参数的话，相当于不要后缀
def switch_ext(file_path, new_ext=""):
    new_ext_path = os.path.join(os.path.dirname(file_path), os.path.splitext(os.path.basename(file_path))[0] + new_ext)
    return new_ext_path


# 判断当前系统是windows还是linux
def get_os():
    os = platform.system()
    if os == "Windows":
        return "w"
    elif os == "Linux":
        return "l"
    else:
        return "o"


# 写一个方法，获取path下，所有的.xxx文件的全路径
def get_all_file_path_list(path, ext):
    # 创建一个list对象all_file_list，用来装所有ext后缀文件的全路径
    all_file_list = []
    # 遍历当前path下的全部文件和文件夹对象temp
    temp = os.walk(path)
    # 遍历temp的全部文件夹
    for i in temp:
        # 遍历当前文件夹下的全部文件
        for file in i[2]:
            # 如果当前文件的后缀为ext
            if file.endswith(ext):
                # 获取当前文件的全路径
                cur_full_path = os.path.join(i[0], file)
                # 添加到all_file_list里
                all_file_list.append(cur_full_path)
    # 返回all_file_list
    return all_file_list


# 这个方法只能输入一个正确的路径
def input_path():
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在
        if os.path.exists(path):
            # 存在的情况，返回此路径，循环结束
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的路径!")
        time.sleep(1)
        print("请重新输入")


def input_path_dir_file():

    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        
        # 如果是直接回车的话，那么就是输入的路径，并且是当前路径
        if path == '':
            path = os.getcwd()
            break
        else:
        # 没有直接回车的情况

            # 在此确认，输入的路径是否存在
            if os.path.exists(path):
                # 存在的情况，返回此路径，循环结束
                break
            # 不存在的情况，提示让用户重新输入
            print("请输入一个正确的路径!")
            time.sleep(1)
            print("请重新输入")
    
    return path




# 这个方法只能输入一个正确的folder path
def input_folder():
    print('直接回车，就是当前，dos的路径')
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        
        if path == '':
            path = os.getcwd()

        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在 并且是一个folder
        if os.path.exists(path) and os.path.isdir(path):
            # 存在并是个folder的情况，返回此路径，循环结束
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的folder路径!")
        time.sleep(1)
        print("请重新输入")

def input_folder_or_none():
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在 并且是一个folder
        if os.path.exists(path) and os.path.isdir(path) or path == "":
            # 存在并是个folder的情况，返回此路径，循环结束
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的folder路径!")
        time.sleep(1)
        print("请重新输入")


# 这个方法只能输入一个正确的file path
def input_file():
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        if path.startswith('"'):
            path = path.replace('"','')

        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在 并且是一个file
        if os.path.exists(path) and os.path.isfile(path):
            # 存在并且是一个file的情况，返回此路径，循环结束
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的file路径!")
        time.sleep(1)
        print("请重新输入")


# 这个方法只能输入一个正确的excel path
def input_excel():
    # 不主动退出循环
    while True:
        # 让用户输入或者拖取一个路径
        path = input()
        # 如果是拖取的情况(此时是'打头的)
        if path.startswith("'"):
            # 去掉'号，变为一个python可以读的路径
            path = path.split("'")[1]
        # 在此确认，输入的路径是否存在，并且一个file，而且还是一个excel file
        if os.path.exists(path) and os.path.isfile(path) and path.endswith(".xlsx"):
            # 如果是：返回路径
            return path
        # 不存在的情况，提示让用户重新输入
        print("请输入一个正确的excel文件路径!")
        time.sleep(1)
        print("请重新输入")


# 给定一个list的集合，里面放入只能输入的参数
def input_params(check_collection=[]):
    while True:
        select = input()
        if select in check_collection:
            return select
        print("输入的参数不匹配！")
        time.sleep(1)
        print("请重新输入")


# 这个方法只能输入数字
def input_digital():
    while True:
        select = input()
        if select.isdigit():
            return select
        print("请输入数字～～")
        time.sleep(1)
        print("请重新输入")


# 这个方法不能输入什么都不输入
def input_not_none():
    while True:
        select = input()
        if select != "":
            return select
        print("不能什么都不输入！！")
        time.sleep(1)
        print("请重新输入")


# 这个方法只能输入数字或者空
def input_digital_or_none():
    while True:
        select = input()
        if select.isdigit() or select == "":
            return select
        print("请输入数字～～")
        time.sleep(1)
        print("请重新输入")


# 执行batch的方法
def run_shell(cmd_list, if_thread=False):
    if isinstance(cmd_list, str):
        cmd = cmd_list
    else:
        cmd = " && ".join(cmd_list)
    rw = None
    if if_thread:
        subprocess.run(cmd, shell=True)
    else:
        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        if p.stdout is not None:
            return_bytes = p.stdout.read()
        else:
            return_bytes = b""
        rw = str(return_bytes)

    return rw


# root并且remount手机
def root_and_remount():
    ROOT = "adb root"
    REMOUNT = "adb remount"
    REMOUNT_SUCCESS_FLAG = "remount succeeded"
    REBOOT_REMINDER = "Now reboot your device for settings to take effect"

    cmd = "&&".join([ROOT, REMOUNT])
    root_remount_rw = run_shell(cmd)
    print("root_remount_rw:" + str(root_remount_rw))
    if REBOOT_REMINDER in str(root_remount_rw):
        print("好像是刷机后的第一次root and remount，要不要重启？")
        print("enter.重启\t enter以外的任意键.不重启")
        select_reboot = input()
        if select_reboot == "":
            reboot_for_wait()
            root_and_remount()
        else:
            print("root remount失败，正在退出程序")
            sys.exit(3)
    else:
        if REMOUNT_SUCCESS_FLAG in str(root_remount_rw):
            print("root和remount成功，adb可以使用")
        else:
            print("root remount失败，正在退出程序")
            sys.exit(3)


def reboot_for_wait():
    run_shell("adb reboot")
    while True:
        un_connect_words = "no devices/emulators found"
        rw = run_shell("adb root")
        if not un_connect_words in str(rw):
            break
        print("didn't connected")
        time.sleep(10)
    print("phone is wake up!")
    time.sleep(20)


# 获取当前时间
def get_time_now():
    formate_words = "%Y%m%d_%H%M%S"
    date_time_now = datetime.datetime.now().strftime(formate_words)
    return date_time_now


# 获取当前电脑的ip
def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()

    return ip


def exe_frame(method,script_name,rounds,times):

    print(f'{script_name}脚本开始运行！')
    time_start = time.perf_counter()

    # 执行方法代码
    method(rounds,times)

    print(f'{script_name}自动化脚本运行结束')
    time_end = time.perf_counter()

    time_cost = time_end - time_start
    cost_min = int(time_cost//60)
    cost_sec = time_cost % 60

    avg_cost_min = int(cost_min / times)
    avg_cost_sec = cost_sec % times
    print(f'花费时间：{cost_min}分{cost_sec:.2f}秒')
    print(f'平均每个rounds的时间：{avg_cost_min}分{avg_cost_sec:.2f}秒')


class adb_logcat_thread(threading.Thread):

    def __init__(self, path="/home/managelocal/Desktop", file="logcat.txt"):
        threading.Thread.__init__(self)
        self.path = path
        self.file = file

    def run(self) -> None:
        self.__run_adb_logcat()

    def __run_adb_logcat(self):
        clean_adb_logcat = "adb logcat -c"
        adb_logcat = "adb logcat > %s/%s" % (self.path, self.file)
        cmd = "&&".join([clean_adb_logcat, adb_logcat])
        print(cmd)
        run_shell(cmd)

    def stop_logcat(self):
        cur_port = self.__get_cur_logcat_port()
        cmd_kill_adb_logcat = "adb shell kill %s" % cur_port
        kill_result = run_shell(cmd_kill_adb_logcat)
        ret_num = -1
        if kill_result == "b''":
            ret_num = 0
        elif "No such process" in str(kill_result):
            ret_num = 1

        return ret_num

    def __get_cur_logcat_port(self):
        cmd = "adb shell ps -A | grep logcat"

        logcat_result = run_shell(cmd)
        return_words = str(logcat_result).split(" ")
        cur_port = None
        for cur_element in return_words:
            if cur_element.isdigit():
                cur_port = cur_element
                break
        return cur_port





# --------------------------------------fgo---------------------------------------------


def run_shell_cmn(x_location,y_location,delay):
    cmd = f'adb shell input tap {x_location} {y_location}'
    run_shell(cmd)
    time.sleep(delay)


def servant_buff(servant_num,skill_num,delay=3.5):
    # 因为3个从者的9个技能都是在一个y轴上的，都是700，所以只需要保留x坐标就行了
    y_location = 700
    servant_buff_x_location = [
        ['这个主要是为了让num不是从0开始。。。。'],
        [0,270,360,480],
        [0,660,760,880],
        [0,1070,1160,1270]
    ]
    x_location = servant_buff_x_location[servant_num][skill_num]

    run_shell_cmn(x_location,y_location,delay)


def servant_buff_target(servant_num,skill_num,target_num,delay=3.5):
    # 点击的哪个技能会有目标指向？
    servant_buff(servant_num,skill_num,delay=1)

    y_location = 500
    servant_target_x_location = [
        0,
        800,
        1150,
        1550
    ]
    x_location = servant_target_x_location[target_num]

    run_shell_cmn(x_location,y_location,delay)
    

def master_buff(buff_number,delay=3.5):

    y_location = 400
    # 御主技能的菱形按钮是2040
    master_buff_x_location = [2040,1685,1785,1885]
    run_shell_cmn(master_buff_x_location[0],y_location,.8)

    x_location = master_buff_x_location[buff_number]
    run_shell_cmn(x_location,y_location,delay)
    

def master_buff_target(buff_number,target_num,delay=3.5):
    # 点击的哪个技能会有目标指向？
    master_buff(buff_number,1)

    y_location = 500
    servant_target_x_location = [
        0,
        800,
        1150,
        1550
    ]
    x_location = servant_target_x_location[target_num]

    run_shell_cmn(x_location,y_location,delay)


def master_buff_sub_off(qian3,hou3):
    
    master_buff(3,.8)
    y_location = 450
    servant_6_location = [
        # 为了从1开始
        0,
        # servant 1
        540,
        # servant 2
        780,
        # servant 3
        1040,
        # servant 4
        1260,
        # servant 5
        1560,
        # servant 6
        1800
    ]

    run_shell_cmn(servant_6_location[qian3],y_location,.8)
    run_shell_cmn(servant_6_location[hou3],y_location,.8)

    sub_off_button = (1150,780)
    run_shell_cmn(sub_off_button[0],sub_off_button[1],4.5)


def push_attack_button():
    # 按攻击按钮
    attck_location = (1950,750)
    run_shell_cmn(attck_location[0],attck_location[1],.8)


def baoju_card(baoju_num,delay=.8):

    y_location = 240
    baoju_x_location = [0,870,1160,1470]
    x_locaiton = baoju_x_location[baoju_num]
    run_shell_cmn(x_locaiton,y_location,delay)


def rbg_card(card_num,delay=.8):

    y_location = 600

    cards_x_location = [
        0,
        # 第1张卡
        520,
        # 第2张卡
        870,
        # 第3张卡
        1160,
        # 第4张卡
        1470,
        # 第5张卡
        1820
    ]

    # 点击那5张攻击的红蓝绿卡
    run_shell_cmn(cards_x_location[card_num],y_location,delay)


def score_screen(if_continue):
    
    # 点击下一步，如果升级的画，也点这个就行，10秒循环点
    next_step_location = (1732,820)

    for i in range(13):
        run_shell_cmn(next_step_location[0],next_step_location[1],1)

    if if_continue:
    # 连续出击的情况
        # 点击连续出击
        continue_attack_button = (1420,740)
        run_shell_cmn(continue_attack_button[0],continue_attack_button[1],2)

        # 选择助战，就选当前画面的第一个
        zhuzhan = (1100,350)
        run_shell_cmn(zhuzhan[0],zhuzhan[1],10)
    else:
    # 回到home的情况
        # 点击关闭
        close_button = (900,740)
        run_shell_cmn(close_button[0],close_button[1],.5)


def main_method(rounds,rounds_name):

    print(f'请输入一个这次脚本使用多少体力？，如果直接回车，不算体力了，默认循环7次。')
    Energys_use = input()

    if Energys_use == '':
        times = 7
    elif not Energys_use.isdigit():
        print('您输入的不是数字，程序结束')
        sys.exit(1)
    else:
        print(f'此本单次的体力消耗是多少？')
        count_total = common.input_digital()
        
        if count_total == '' or not count_total.isdigit():
            print('您输入的不是数字，程序结束')
            sys.exit(1)

        times = int(int(Energys_use) / int(count_total))

    exe_frame(fgo_main_method,rounds_name,rounds,times)


def fgo_main_method(rounds,times):

    if_continue = True
    for i in range(times):
        # 如果只跑一次的话，不需要循环
        if times == 1:
            if_continue = False
        
        # 最后一个循环的时候，也不需要再来了
        if i == times - 1:
            if_continue = False
        
        print(f'现在进度：{i+1}/{times}')
            
        rounds(if_continue)

    
