import os
import shutil
import fitz

import openpyxl
from pdfrw import PdfReader
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

import TOOLS_COMMON as common

default_merge_page_count = 6

def main():
    print("请输入docx folder的路径：")
    folder_path = common.input_folder()

    print("请输入接近多少页数合并为一个file，直接回车的话，默认为"+str(default_merge_page_count)+"页")
    close_to = common.input_digital_or_none()
    if close_to == "":
        close_to = default_merge_page_count
    else:
        close_to = int(close_to)

    print("请输入新的pdf folder名字：直接回车的话，默认为temp_folder")
    package_name = input()

    print("请输入合并后的pdf folder名字：直接回车的话，默认为merged")
    output_folder_name = input()
    if output_folder_name == "":
        output_folder_name = 'merged'

    # 先将folder里所有的docx全部转成pdf
    print("执行进度：1.先将folder里所有的docx全部转成pdf")
    print()
    common.word_to_pdf(folder_path)

    # 将pdf全部打包，放入新的folder里，获取新folder的全路径
    print("执行进度：2.将pdf全部打包，放入新的folder里，获取新folder的全路径")
    print()
    new_folder_path = package_file(folder_path,package_name)
    # new_folder_path = r"D:\OneDrive\日本語\EP04_resource\test\temp_folder"

    print("执行进度：3.将新folder里的pdf文件以和合并页数合并，如果直接回车，我默认为新的合并pdf folder的名字是merged:")
    output_data = _merge_pdf(new_folder_path,output_folder_name,close_to)
    print(output_data)
    print()

    # 统计所有的pdf名字和页数，放入新建的excel里
    print("执行进度：4.统计所有的pdf名字和页数，获取数据")
    print()
    input_dict,data_merged = get_single_count_pdf(new_folder_path, output_data)

    print("==========================")
    print(data_merged)
    print("==========================")
    # 将统计出来的数据输入到一个excel文件里
    print("执行进度：5.将统计出来的数据输入到一个excel文件里")
    print()
    caculate_data_to_excel(input_dict, os.path.join(new_folder_path,output_folder_name),data_merged)

def _merge_pdf(path,output_folder_name,close_to):
    scene_name_list = []
    scene_page_count_list = []
    output_data = {}

    files = os.listdir(path)
    print("初始化")
    for file in files:
        if file.endswith(".pdf"):
            scene_name_list.append(file.replace(".pdf",""))
            pdf_file_path = os.path.join(path,file)
            scene_page_count_list.append(_get_pdf_file_page_count(pdf_file_path))

    print(scene_name_list)
    print(scene_page_count_list)

    print("============================")
    print("开始解析")
    while True:
        if len(scene_name_list) <= 0:
            break
        page_count = 0

        scene_data = [0]
        for i in range(0, len(scene_name_list)):
            last_page_count = page_count
            page_count += scene_page_count_list[i]

            if page_count >= close_to or i == len(scene_name_list) - 1:
                aaa = page_count - close_to
                bbb = close_to - last_page_count
                if aaa > bbb:
                    scene_data[0] = last_page_count
                    delete_index = i
                    name_index = i - 1
                else:
                    scene_data[0] = page_count
                    scene_data.append(scene_name_list[i])
                    delete_index = i + 1
                    name_index = i

                if name_index != 0:
                    ten_pdf_name = scene_name_list[0].split(".")[0] + "~" + scene_name_list[name_index].split(".")[0]
                else:
                    ten_pdf_name = scene_name_list[name_index].split(".")[0]

                scene_name_list = scene_name_list[delete_index:]
                scene_page_count_list = scene_page_count_list[delete_index:]

                output_data[ten_pdf_name] = scene_data
                break
            else:
                scene_data.append(scene_name_list[i])

            print(scene_name_list)
            print(scene_page_count_list)
            print("output_data:++++++++++++++")
            print(output_data)

    print("开始合成")
    _merge_pdf_to_1_file(output_data,path,output_folder_name)
    print("合成完毕！")

    return output_data



def _merge_pdf_to_1_file(output_data,path,output_folder_name):
    output_foler_path = os.path.join(path,output_folder_name)
    if not os.path.exists(output_foler_path):
        os.mkdir(output_foler_path)
    for key,value in output_data.items():
        output_file_name = os.path.join(output_foler_path,key+".pdf")
        print(value[0])
        merged_file_list = []
        for i in range(1,len(value)):
            merged_file_list.append(os.path.join(path,value[i]+".pdf"))

        common.merge_pdfs(merged_file_list,output_file_name)
        print(output_file_name + "合并完成！！！")

def _get_pdf_file_page_count(pdf_file_path):
    pdf_document = fitz.open(pdf_file_path)
    return pdf_document.page_count

def package_file(folder_path, package_name):
    if package_name == "":
        package_name = "temp_folder"

    # 新的c folder的路径
    c_folder_path = os.path.join(folder_path, package_name)

    # 如果新的c folder的path不存在，建一个
    if not os.path.exists(c_folder_path):
        os.mkdir(c_folder_path)

    # 移动所有pdf文件到新的folder里
    for file in os.listdir(folder_path):
        if file.endswith(".pdf"):
            cur_file_path = os.path.join(folder_path, file)
            shutil.move(cur_file_path, c_folder_path)

    return c_folder_path

def get_single_count_pdf(folder_path,output_data):
    input_dict = {}
    for file in os.listdir(folder_path):
            if file.endswith(".pdf"):
                # 找到当前pdf的全路径
                cur_pdf_path = os.path.join(folder_path, file)
                # 找出当前pdf有多少页
                pages = len(PdfReader(cur_pdf_path).pages)
                # 找出当前的数据，以键值对的形式：filename：pages，放入全部data里
                input_dict[os.path.basename(cur_pdf_path)] = pages

    data_merged = []
    for key,value in output_data.items():
        data_key = value[0]
        data_value = len(value) - 1
        data_tuble = (data_key,data_value)
        data_merged.append(data_tuble)

    return input_dict,data_merged


def _write_to_merged_cells(ws: Worksheet, data: dict, start_col = 3, start_row = 2):
    """
    Write values to merged cells in an Excel sheet and center the text.

    :param ws: The worksheet object from openpyxl.
    :param data: A dictionary where keys are the values to write and the values are the number of rows to merge.
    :param start_col: The starting column index (1-based).
    :param start_row: The starting row index (1-based).
    """
    current_row = start_row
    start_col_letter = get_column_letter(start_col)

    for data_tuble in data:
        end_row = current_row + data_tuble[1] - 1
        cell_range = f"{start_col_letter}{current_row}:{start_col_letter}{end_row}"
        ws.merge_cells(cell_range)
        cell = ws[f"{start_col_letter}{current_row}"]
        cell.value = data_tuble[0]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row = end_row + 1

def caculate_data_to_excel(input_dict,new_folder_path,data_merged):
    excel_path = os.path.join(new_folder_path,"result.xlsx")

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 为工作表添加标题（可选）
    sheet['A1'] = 'name'
    sheet['B1'] = 'pages'
    sheet['C1'] = 'merged'

    # 填充数据
    for i, (key, value) in enumerate(input_dict.items(), start=2):
        sheet[f'A{i}'] = key
        sheet[f'B{i}'] = value

    _write_to_merged_cells(sheet, data_merged)

    # 定义全局字体格式
    common.set_excel_font_style(sheet)

    # 保存工作簿
    workbook.save(excel_path)



if __name__ == '__main__':
    main()
