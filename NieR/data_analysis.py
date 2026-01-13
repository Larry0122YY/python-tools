import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
import sys
from matplotlib.font_manager import FontProperties

import TOOLS_COMMON as common


def main():
    print("请将excel文件拖入这里：(目前的版本只会读取Sheet1的数据)")
    path = common.input_excel()

    titles, datas = get_Datas(path)

    print("默认是圆饼图(直接回车)，如果您想看柱状图，请输入回车以外：")
    形状 = input()
    zhutu = False;
    if 形状 == "":
        zhutu = True;

    if zhutu:
        柱状图(titles, datas)
    else:
        圆饼图(titles, datas)

    print("程序运行完毕~！")


def get_Datas(path):
    titles = []
    datas = []

    book = openpyxl.load_workbook(path)
    sheet = book["Sheet1"]
    cur_max_column = sheet.max_column
    cur_max_row = sheet.max_row

    # 获取所有的title名
    for i in range(1, cur_max_column + 1):
        titles.append(sheet.cell(row=1, column=i).value)
    
    # 获取每个title的所有钱
    for i in range(1, cur_max_column + 1):
        one_title_count = 0
        for j in range(2, cur_max_row + 1):
            cur_value = sheet.cell(row=j, column=i).value
            if cur_value is None:
                break
            # 判断 cur_value 是否为数字类型
            if not isinstance(cur_value, (int, float)):
                # 如果不是数字类型，打印日志并退出程序
                print(f"错误：单元格 Sheet1!{get_column_letter(i)}{j} 的值不是数字类型")
                print(f"  当前值：{cur_value}")
                print(f"  当前类型：{type(cur_value).__name__}")
                print(f"  期望类型：int 或 float")
                print("程序终止执行")
                sys.exit(1)
            one_title_count += int(cur_value)
        datas.append(one_title_count)

    return titles, datas


def 柱状图(categories, values):
    # 创建柱状图
    plt.bar(categories, values)

    # 添加标题和标签
    plt.title('expenses')
    plt.xlabel('items')
    plt.ylabel('costs')

    # 显示图表
    plt.show()


def 圆饼图(labels, sizes):
    import matplotlib.pyplot as plt

    # 创建饼图
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)

    # 添加标题
    plt.title('expenses')

    # 显示图表
    plt.show()


if __name__ == '__main__':
    main()
    # get_Datas("C:\\Users\\Administrator\\Desktop\\test.xlsx")
